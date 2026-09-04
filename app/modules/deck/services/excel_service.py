from __future__ import annotations
import pandas as pd
from typing import List, Dict, Any, Tuple, Optional, Union, Set
from io import BytesIO
import json
import re

# MindStack COLUMN_ALIASES
COLUMN_ALIASES = {
    # Common
    'item_id': {'item_id', 'id', 'id câu hỏi', 'id item'},
    'order_in_container': {'order', 'stt', 'order_in_container', 'thứ tự', 'sắp xếp'},

    # Flashcard
    'front': {'front', 'mặt trước', 'mat truoc', 'term', 'từ vựng', 'tu vung', 'từ', 'text 1', 'question'},
    'back': {'back', 'mặt sau', 'mat sau', 'definition', 'định nghĩa', 'dinh nghia', 'nghĩa', 'nghia', 'answer', 'text 2'},
    'front_img': {'front_img', 'ảnh mặt trước', 'anh mat truoc', 'front image', 'image 1', 'image'},
    'back_img': {'back_img', 'ảnh mặt sau', 'anh mat sau', 'back image', 'image 2'},
    'front_audio_url': {'front_audio_url', 'audio mặt trước', 'audio mat truoc', 'audio 1', 'audio front', 'audio'},
    'back_audio_url': {'back_audio_url', 'audio mặt sau', 'audio mat sau', 'audio 2', 'audio back'},
    'front_audio_content': {'front_audio_content', 'văn bản audio mặt trước', 'front audio content'},
    'back_audio_content': {'back_audio_content', 'văn bản audio mặt sau', 'back audio content'},

    # Quiz
    'question': {'question', 'câu hỏi', 'cau hoi', 'nội dung câu hỏi', 'noidung', 'content', 'text 1', 'q'},
    'correct_answer': {'correct_answer', 'correct answer', 'đáp án đúng', 'dap an dung', 'đáp án', 'dap an', 'answer', 'ans', 'correct', 'key', 'result'},
    'explanation': {'explanation', 'giải thích', 'giai thich', 'lời giải', 'loi giai', 'explain', 'suggest', 'hint', 'gợi ý', 'goi y', 'guidance'},
    'option_a': {'option_a', 'option a', 'lựa chọn a', 'lua chon a', 'a', 'đáp án a', 'dap an a', 'choice a'},
    'option_b': {'option_b', 'option b', 'lựa chọn b', 'lua chon b', 'b', 'đáp án b', 'dap an b', 'choice b'},
    'option_c': {'option_c', 'option c', 'lựa chọn c', 'lua chon c', 'c', 'đáp án c', 'dap an c', 'choice c'},
    'option_d': {'option_d', 'option d', 'lựa chọn d', 'lua chon d', 'd', 'đáp án d', 'dap an d', 'choice d'},
    'pre_question_text': {'pre_question_text', 'pre question', 'đoạn văn trước', 'doan van truoc', 'context', 'bối cảnh'},
    
    # AI
    'ai_explanation': {'ai_explanation', 'ai giải thích', 'ai giai thich'},
    'ai_prompt': {'ai_prompt', 'ai prompt', 'prompt'},

    # Custom Game Modes & Data
    'other_content': {'other_content', 'other content', 'nội dung khác', 'noi dung khac', 'custom_data', 'others', 'other_data'}
}

def normalize_column_headers(columns: List[str]) -> Dict[str, str]:
    """
    Map raw column names to standardized field names based on aliases.
    Exclusively maps each standard name to at most one raw column.
    """
    mapping = {}
    used_standards = set()
    
    # Pass 1: Case-insensitive Exact Match (Priority)
    for col in columns:
        clean_col = str(col).strip().lower()
        if clean_col in COLUMN_ALIASES and clean_col not in used_standards:
            mapping[col] = clean_col
            used_standards.add(clean_col)
            
    # Pass 2: Alias Match for remaining columns
    for col in columns:
        if col in mapping:
            continue
            
        clean_col = str(col).strip().lower()
        for standard, aliases in COLUMN_ALIASES.items():
            if standard in used_standards:
                continue
            if clean_col in aliases:
                mapping[col] = standard
                used_standards.add(standard)
                break
                
    # Unmapped columns map to lowered original name
    for col in columns:
        if col not in mapping:
            mapping[col] = str(col).strip().lower()
            
    return mapping


def tokenize_formula(formula: str, current_row: int, col_letter_to_name: Dict[str, str]) -> str:
    """
    Converts physical cell references in formula (e.g. A2, B2) to semantic tokens {{col_name}}.
    Safely preserves strings inside quotes (e.g. "A2").
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula

    pattern = re.compile(r'(?<![A-Za-z0-9_])(\$?)([A-Za-z]{1,3})(\$?)(\d+)(?![A-Za-z0-9_])')
    parts = re.split(r'("(?:[^"]|"")*")', formula)

    def replace_in_part(text: str) -> str:
        def replacer(match):
            abs_col, col_letter, abs_row, row_num = match.groups()
            row_int = int(row_num)
            col_upper = col_letter.upper()

            if col_upper in col_letter_to_name:
                col_name = col_letter_to_name[col_upper]
                if row_int == current_row:
                    return f"{{{{{col_name}}}}}"
                else:
                    return f"{{{{{col_name}:{row_num}}}}}"
            return match.group(0)

        return pattern.sub(replacer, text)

    result_parts = []
    for p in parts:
        if p.startswith('"') and p.endswith('"'):
            result_parts.append(p)
        else:
            result_parts.append(replace_in_part(p))

    return "".join(result_parts)


def transpile_formula(tokenized_formula: str, target_row: int, col_name_to_letter: Dict[str, str]) -> str:
    """
    Converts semantic tokens {{col_name}} to physical cell coordinates (e.g. B2, K2) based on exported column layout.
    """
    if not isinstance(tokenized_formula, str) or not tokenized_formula.startswith("="):
        return tokenized_formula

    pattern = re.compile(r'\{\{([a-zA-Z0-9_]+)(?::(\d+))?\}\}')

    def replacer(match):
        col_name, explicit_row = match.groups()
        target_col_letter = col_name_to_letter.get(col_name)
        if not target_col_letter:
            return f"#{col_name}!"
        row_str = explicit_row if explicit_row else str(target_row)
        return f"{target_col_letter}{row_str}"

    return pattern.sub(replacer, tokenized_formula)


class ExcelDeckService:
    @staticmethod
    def parse_deck_excel(file_content: bytes) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
        """
        Parses an Excel file matching MindStack's structure with 'Info' and 'Data' sheets.
        Handles both Quiz Format (MCQ) and Flashcard/Vocab Format (front/back).
        """
        try:
            print(f"DEBUG: Loading Excel file into pandas...")
            excel_file = pd.ExcelFile(BytesIO(file_content))
            print(f"DEBUG: Sheet names found: {excel_file.sheet_names}")
        except Exception as e:
            print(f"CRITICAL: Excel loading error: {e}")
            return {}, []

        # 1. Parse 'Info' sheet for metadata
        metadata: Dict[str, Any] = {
            "title": "Imported Deck",
            "description": "",
            "category": "General",
            "cover_image": "",
            "instruction": "",
            "is_public": True,
            "time_limit": 0,
            "tags": [],
            "practice_settings": {},
            "collaborators": []
        }
        
        if "Info" in excel_file.sheet_names:
            print("DEBUG: Parsing 'Info' sheet...")
            df_info = excel_file.parse("Info")
            # Normalize Info sheet columns
            df_info.columns = [str(c).strip().lower() for c in df_info.columns]
            
            if "key" in df_info.columns and "value" in df_info.columns:
                for _, row in df_info.iterrows():
                    key = str(row.get("key", "")).strip().lower()
                    value = str(row.get("value", "")).strip()
                    if not value or value.lower() == "nan": continue
                    
                    if key == "title": metadata["title"] = value
                    elif key == "description": metadata["description"] = value
                    elif key == "category": metadata["category"] = value
                    elif key == "cover_image": metadata["cover_image"] = value
                    elif key == "instruction": metadata["instruction"] = value
                    elif key in ("is_public", "public", "công khai"):
                        metadata["is_public"] = value.lower() in ("true", "1", "yes", "y", "công khai", "public")
                    elif key == "tags": metadata["tags"] = [t.strip() for t in value.split(",") if t.strip()]
                    elif key == "practice_settings":
                        try:
                            parsed = json.loads(value)
                            if isinstance(parsed, dict):
                                if "practice_settings" not in metadata or not isinstance(metadata["practice_settings"], dict):
                                    metadata["practice_settings"] = {}
                                metadata["practice_settings"].update(parsed)
                        except:
                            pass
                    elif key in ("study_defaults", "cài đặt học mặc định", "cài đặt học"):
                        try:
                            parsed = json.loads(value)
                            if isinstance(parsed, dict):
                                if "practice_settings" not in metadata or not isinstance(metadata["practice_settings"], dict):
                                    metadata["practice_settings"] = {}
                                if "study_defaults" not in metadata["practice_settings"] or not isinstance(metadata["practice_settings"]["study_defaults"], dict):
                                    metadata["practice_settings"]["study_defaults"] = {}
                                from app.modules.deck.utils import normalize_study_setting_value, STUDY_SETTINGS_KEYS
                                for sk, sv in parsed.items():
                                    norm_key = sk[6:] if sk.startswith("study_") else sk
                                    if norm_key in STUDY_SETTINGS_KEYS:
                                        norm_val = normalize_study_setting_value(norm_key, sv)
                                        if norm_val is not None:
                                            metadata["practice_settings"]["study_defaults"][norm_key] = norm_val
                        except:
                            pass
                    elif key in (
                        "study_autoplay_audio", "autoplay_audio", "tự động phát âm thanh", "phát âm thanh",
                        "study_show_images", "show_images", "hiển thị hình ảnh", "hình ảnh",
                        "study_learning_mode", "learning_mode", "chế độ học mặc định", "chế độ học",
                        "study_random_enabled", "random_enabled", "xáo trộn thẻ", "random thẻ",
                        "study_sfx_enabled", "sfx_enabled", "âm thanh hiệu ứng", "hiệu ứng âm thanh",
                        "study_haptic_enabled", "haptic_enabled", "rung phản hồi", "haptic",
                        "study_quick_learn_enabled", "quick_learn_enabled", "học nhanh", "chế độ học nhanh",
                        "study_show_fsrs", "show_fsrs", "hiển thị nút fsrs", "nút đánh giá fsrs"
                    ):
                        from app.modules.deck.utils import normalize_study_setting_value
                        clean_key = key[6:] if key.startswith("study_") else key
                        key_aliases = {
                            "tự động phát âm thanh": "autoplay_audio",
                            "phát âm thanh": "autoplay_audio",
                            "hiển thị hình ảnh": "show_images",
                            "hình ảnh": "show_images",
                            "chế độ học mặc định": "learning_mode",
                            "chế độ học": "learning_mode",
                            "xáo trộn thẻ": "random_enabled",
                            "random thẻ": "random_enabled",
                            "âm thanh hiệu ứng": "sfx_enabled",
                            "hiệu ứng âm thanh": "sfx_enabled",
                            "rung phản hồi": "haptic_enabled",
                            "haptic": "haptic_enabled",
                            "học nhanh": "quick_learn_enabled",
                            "chế độ học nhanh": "quick_learn_enabled",
                            "hiển thị nút fsrs": "show_fsrs",
                            "nút đánh giá fsrs": "show_fsrs",
                        }
                        target_key = key_aliases.get(clean_key, clean_key)
                        norm_val = normalize_study_setting_value(target_key, value)
                        if norm_val is not None:
                            if "practice_settings" not in metadata or not isinstance(metadata["practice_settings"], dict):
                                metadata["practice_settings"] = {}
                            if "study_defaults" not in metadata["practice_settings"] or not isinstance(metadata["practice_settings"]["study_defaults"], dict):
                                metadata["practice_settings"]["study_defaults"] = {}
                            metadata["practice_settings"]["study_defaults"][target_key] = norm_val
                    elif key in ("custom_columns", "insight_columns", "column_order", "study_modes", "default_mode", "ai_prompts", "audio_pairs", "front_audio_config", "back_audio_config", "mcq", "typing", "listening"):
                        try:
                            parsed_val = json.loads(value)
                        except:
                            if key in ("custom_columns", "insight_columns", "column_order", "study_modes"):
                                parsed_val = [t.strip() for t in value.split(",") if t.strip()]
                            else:
                                parsed_val = value
                        
                        if "practice_settings" not in metadata or not isinstance(metadata["practice_settings"], dict):
                            metadata["practice_settings"] = {}
                        metadata["practice_settings"][key] = parsed_val
                    elif key in ("ai_prompt", "ai_prompt_explanation", "prompt_explanation"):
                        metadata["ai_prompt"] = value
                        if "practice_settings" not in metadata: metadata["practice_settings"] = {}
                        metadata["practice_settings"]["ai_prompt"] = value
                    elif key in ("ai_prompt_hint", "prompt_hint"):
                        metadata["ai_prompt_hint"] = value
                        if "practice_settings" not in metadata: metadata["practice_settings"] = {}
                        metadata["practice_settings"]["ai_prompt_hint"] = value
                    elif key in ("ai_prompt_mnemonic", "prompt_mnemonic"):
                        metadata["ai_prompt_mnemonic"] = value
                        if "practice_settings" not in metadata: metadata["practice_settings"] = {}
                        metadata["practice_settings"]["ai_prompt_mnemonic"] = value
                    elif key == "time_limit": 
                        try: metadata["time_limit"] = int(float(value))
                        except: pass
                    elif key in ("active_pairs", "practice_pairs", "cấu hình luyện tập", "cặp câu hỏi luyện tập", "luyện tập", "mcq_active_pairs"):
                        pairs = []
                        parts = re.split(r'[,;]+', value)
                        for part in parts:
                            part = part.strip()
                            if not part: continue
                            subparts = re.split(r'->|-|:', part)
                            if len(subparts) >= 2:
                                q_col = subparts[0].strip().lower()
                                a_col = subparts[1].strip().lower()
                                if q_col and a_col:
                                    pairs.append({"q": q_col, "a": a_col})
                        if pairs:
                            if "practice_settings" not in metadata:
                                metadata["practice_settings"] = {}
                            if "mcq" not in metadata["practice_settings"] or not isinstance(metadata["practice_settings"]["mcq"], dict):
                                metadata["practice_settings"]["mcq"] = {}
                            metadata["practice_settings"]["mcq"]["active_pairs"] = pairs
                            metadata["practice_settings"]["active_pairs"] = pairs
                    elif key in ("typing_active_pairs", "typing_pairs"):
                        pairs = []
                        parts = re.split(r'[,;]+', value)
                        for part in parts:
                            part = part.strip()
                            if not part: continue
                            subparts = re.split(r'->|-|:', part)
                            if len(subparts) >= 2:
                                q_col = subparts[0].strip().lower()
                                a_col = subparts[1].strip().lower()
                                if q_col and a_col:
                                    pairs.append({"q": q_col, "a": a_col})
                        if pairs:
                            if "practice_settings" not in metadata: metadata["practice_settings"] = {}
                            if "typing" not in metadata["practice_settings"] or not isinstance(metadata["practice_settings"]["typing"], dict):
                                metadata["practice_settings"]["typing"] = {}
                            metadata["practice_settings"]["typing"]["active_pairs"] = pairs
                    elif key in ("listening_active_pairs", "listening_pairs"):
                        pairs = []
                        parts = re.split(r'[,;]+', value)
                        for part in parts:
                            part = part.strip()
                            if not part: continue
                            subparts = re.split(r'->|-|:', part)
                            if len(subparts) >= 2:
                                q_col = subparts[0].strip().lower()
                                a_col = subparts[1].strip().lower()
                                if q_col and a_col:
                                    pairs.append({"q": q_col, "a": a_col})
                        if pairs:
                            if "practice_settings" not in metadata: metadata["practice_settings"] = {}
                            if "listening" not in metadata["practice_settings"] or not isinstance(metadata["practice_settings"]["listening"], dict):
                                metadata["practice_settings"]["listening"] = {}
                            metadata["practice_settings"]["listening"]["active_pairs"] = pairs
                    elif key in ("num_choices", "practice_num_choices", "số lựa chọn", "mcq_num_choices"):
                        try:
                            num = int(float(value))
                            if 3 <= num <= 8:
                                if "practice_settings" not in metadata: metadata["practice_settings"] = {}
                                if "mcq" not in metadata["practice_settings"] or not isinstance(metadata["practice_settings"]["mcq"], dict):
                                    metadata["practice_settings"]["mcq"] = {}
                                metadata["practice_settings"]["mcq"]["num_choices"] = num
                                metadata["practice_settings"]["num_choices"] = num
                        except:
                            pass
                    elif key in ("listening_num_choices", "số lựa chọn nghe"):
                        try:
                            num = int(float(value))
                            if 3 <= num <= 8:
                                if "practice_settings" not in metadata: metadata["practice_settings"] = {}
                                if "listening" not in metadata["practice_settings"] or not isinstance(metadata["practice_settings"]["listening"], dict):
                                    metadata["practice_settings"]["listening"] = {}
                                metadata["practice_settings"]["listening"]["num_choices"] = num
                        except:
                            pass
                    elif key in ("collaborators", "cộng tác viên", "contributors"):
                        collabs = []
                        parts = re.split(r'[,;]+', value)
                        for part in parts:
                            part = part.strip()
                            if not part: continue
                            if ":" in part:
                                uname, role = part.split(":", 1)
                                collabs.append({"username": uname.strip(), "role": role.strip()})
                            else:
                                collabs.append({"username": part.strip(), "role": "editor"})
                        if collabs:
                            metadata["collaborators"] = collabs
        
        try:
            print(f"DEBUG: Metadata extracted: {metadata.get('title', '')}")
        except Exception:
            pass

        # 2. Parse 'Data_Formula' or 'Data' sheet for questions / cards
        questions = []
        if not excel_file.sheet_names:
            return metadata, []
            
        sheet_name = "Data_Formula" if "Data_Formula" in excel_file.sheet_names else ("Data" if "Data" in excel_file.sheet_names else excel_file.sheet_names[0])
        print(f"DEBUG: Parsing '{sheet_name}' sheet...")

        import openpyxl
        from openpyxl.utils import get_column_letter

        openpyxl_parsed = False
        deck_column_formulas = {}

        try:
            wb_raw = openpyxl.load_workbook(BytesIO(file_content), data_only=False)
            wb_val = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
            if sheet_name in wb_raw.sheetnames:
                ws_raw = wb_raw[sheet_name]
                ws_val = wb_val[sheet_name] if (wb_val and sheet_name in wb_val.sheetnames) else ws_raw

                # Read header row (row 1)
                raw_headers = [ws_raw.cell(row=1, column=c).value for c in range(1, ws_raw.max_column + 1)]
                raw_cols = [str(c).strip() for c in raw_headers if c is not None and str(c).strip()]
                mapping = normalize_column_headers(raw_cols)

                col_letter_to_name = {}
                for c_idx, h in enumerate(raw_headers, start=1):
                    if h is not None and str(h).strip():
                        clean_h = str(h).strip()
                        norm_h = mapping.get(clean_h, clean_h.lower())
                        col_letter_to_name[get_column_letter(c_idx)] = norm_h

                for r in range(2, ws_raw.max_row + 1):
                    row_dict = {}
                    row_formulas = {}
                    has_data = False

                    for c_idx in range(1, ws_raw.max_column + 1):
                        col_let = get_column_letter(c_idx)
                        col_name = col_letter_to_name.get(col_let)
                        if not col_name:
                            continue

                        raw_cell = ws_raw.cell(row=r, column=c_idx).value
                        val_cell = ws_val.cell(row=r, column=c_idx).value

                        if isinstance(raw_cell, str) and raw_cell.strip().startswith("="):
                            tokenized = tokenize_formula(raw_cell.strip(), r, col_letter_to_name)
                            row_formulas[col_name] = tokenized
                            if col_name not in deck_column_formulas:
                                deck_column_formulas[col_name] = tokenized
                            cell_value = val_cell if val_cell is not None else raw_cell
                        else:
                            cell_value = val_cell if val_cell is not None else raw_cell

                        if cell_value is not None:
                            s_val = str(cell_value).strip()
                            s_val = s_val.replace('_x000D_', '').replace('_x000d_', '')
                            s_val = s_val.replace('\\r\\n', '\n').replace('\\n', '\n')
                            if s_val and s_val.lower() != "nan":
                                row_dict[col_name] = s_val
                                has_data = True

                    if not has_data:
                        continue

                    front_text = row_dict.get("front") or row_dict.get("question")
                    if not front_text or front_text.lower() == "nan":
                        continue

                    explanation_text = row_dict.get("back") or row_dict.get("correct_answer") or row_dict.get("explanation")

                    options_list = []
                    for opt_key in ['option_a', 'option_b', 'option_c', 'option_d']:
                        opt_val = row_dict.get(opt_key)
                        if opt_val and opt_val.lower() != "nan":
                            is_correct = (opt_val == explanation_text) or (opt_key == 'option_a' and not explanation_text)
                            options_list.append({
                                "content": opt_val,
                                "is_correct": is_correct
                            })

                    others_dict = {}
                    for col, val in row_dict.items():
                        if col not in ('id', 'item_id', 'order_in_container', 'front', 'back', 'question', 'correct_answer', 'explanation', 
                                       'option_a', 'option_b', 'option_c', 'option_d', 'front_audio_content', 'back_audio_content',
                                       'front_audio_url', 'back_audio_url', 'front_img', 'back_img', 'audio', 'image', '_formulas', '_formula'):
                            others_dict[col] = val

                    raw_other_content = row_dict.get("other_content")
                    if raw_other_content and raw_other_content.lower() != "nan":
                        try:
                            parsed_other = json.loads(raw_other_content)
                            if isinstance(parsed_other, dict):
                                others_dict.update(parsed_other)
                        except Exception:
                            others_dict["other_content"] = raw_other_content

                    if row_formulas:
                        others_dict["_formulas"] = row_formulas

                    question_data = {
                        "id": row_dict.get("id") or row_dict.get("item_id") or None,
                        "content": front_text,
                        "explanation": explanation_text,
                        "front_img": row_dict.get("front_img") or row_dict.get("image"),
                        "back_img": row_dict.get("back_img"),
                        "front_audio_url": row_dict.get("front_audio_url") or row_dict.get("audio"),
                        "back_audio_url": row_dict.get("back_audio_url"),
                        "front_audio_content": row_dict.get("front_audio_content"),
                        "back_audio_content": row_dict.get("back_audio_content"),
                        "options": options_list,
                        "others": others_dict
                    }
                    questions.append(question_data)

                openpyxl_parsed = True
        except Exception as e:
            print(f"DEBUG: openpyxl extraction warning: {e}, using pandas fallback")

        # Fallback to pandas if openpyxl parsing was not performed
        if not openpyxl_parsed:
            df_data = excel_file.parse(sheet_name)
            raw_cols = [str(c).strip() for c in df_data.columns]
            mapping = normalize_column_headers(raw_cols)
            df_data.rename(columns=mapping, inplace=True)
            
            for idx, row in df_data.iterrows():
                def get_val(col, default=""):
                    try:
                        val = row.get(col)
                        if pd.notna(val):
                            s_val = str(val).strip()
                            s_val = s_val.replace('_x000D_', '').replace('_x000d_', '')
                            s_val = s_val.replace('\\r\\n', '\n').replace('\\n', '\n')
                            return s_val
                        return default
                    except Exception:
                        return default

                front_text = get_val("front") or get_val("question")
                if not front_text or front_text.lower() == "nan":
                    continue

                explanation_text = get_val("back") or get_val("correct_answer") or get_val("explanation")
                options_list = []
                for opt_key in ['option_a', 'option_b', 'option_c', 'option_d']:
                    opt_val = get_val(opt_key)
                    if opt_val and opt_val.lower() != "nan":
                        is_correct = (opt_val == explanation_text) or (opt_key == 'option_a' and not explanation_text)
                        options_list.append({"content": opt_val, "is_correct": is_correct})

                others_dict = {}
                for col in df_data.columns:
                    if col not in ('id', 'item_id', 'order_in_container', 'front', 'back', 'question', 'correct_answer', 'explanation', 
                                   'option_a', 'option_b', 'option_c', 'option_d', 'front_audio_content', 'back_audio_content',
                                   'front_audio_url', 'back_audio_url', 'front_img', 'back_img', 'audio', 'image', '_formulas', '_formula'):
                        val = get_val(col)
                        if val and val.lower() != "nan":
                            others_dict[col] = val

                raw_other_content = get_val("other_content")
                if raw_other_content and raw_other_content.lower() != "nan":
                    try:
                        parsed_other = json.loads(raw_other_content)
                        if isinstance(parsed_other, dict):
                            others_dict.update(parsed_other)
                    except Exception:
                        others_dict["other_content"] = raw_other_content

                question_data = {
                    "id": get_val("id") or get_val("item_id") or None,
                    "content": front_text,
                    "explanation": explanation_text,
                    "front_img": get_val("front_img") or get_val("image"),
                    "back_img": get_val("back_img"),
                    "front_audio_url": get_val("front_audio_url") or get_val("audio"),
                    "back_audio_url": get_val("back_audio_url"),
                    "front_audio_content": get_val("front_audio_content"),
                    "back_audio_content": get_val("back_audio_content"),
                    "options": options_list,
                    "others": others_dict
                }
                questions.append(question_data)

        if deck_column_formulas:
            if "practice_settings" not in metadata or not isinstance(metadata["practice_settings"], dict):
                metadata["practice_settings"] = {}
            if "column_formulas" not in metadata["practice_settings"]:
                metadata["practice_settings"]["column_formulas"] = {}
            metadata["practice_settings"]["column_formulas"].update(deck_column_formulas)
                    
        return metadata, questions

    @staticmethod
    def export_deck_to_excel(
        deck_title: str, 
        deck_description: str, 
        category_name: str, 
        tags: List[str], 
        practice_settings: Dict[str, Any], 
        cards: List[Any], 
        exclude_ids: bool = False,
        cover_image: str = "",
        instruction: str = "",
        is_public: bool = True,
        time_limit: int = 0,
        ai_prompt: str = "",
        ai_prompt_hint: str = "",
        ai_prompt_mnemonic: str = "",
        collaborators: List[Dict[str, Any]] = None
    ) -> bytes:
        """
        Generates an Excel workbook (bytes) containing comprehensive Info and Data sheets
        for exporting a quiz/deck with 100% settings fidelity.
        """
        output = BytesIO()
        ps = practice_settings or {}
        
        # 1. Prepare Info sheet key-value data covering all 7 Edit Collection tabs
        info_data = [
            {"key": "title", "value": deck_title or ""},
            {"key": "description", "value": deck_description or ""},
            {"key": "category", "value": category_name or "General"},
            {"key": "tags", "value": ", ".join(tags) if tags else ""},
            {"key": "cover_image", "value": cover_image or ""},
            {"key": "instruction", "value": instruction or ""},
            {"key": "is_public", "value": "TRUE" if is_public else "FALSE"},
            {"key": "time_limit", "value": str(time_limit or 0)},
        ]
        
        # Column Manager
        if "column_order" in ps and isinstance(ps["column_order"], list):
            info_data.append({"key": "column_order", "value": ", ".join(ps["column_order"])})
        if "custom_columns" in ps and isinstance(ps["custom_columns"], list):
            info_data.append({"key": "custom_columns", "value": ", ".join(ps["custom_columns"])})
        if "insight_columns" in ps and isinstance(ps["insight_columns"], list):
            info_data.append({"key": "insight_columns", "value": ", ".join(ps["insight_columns"])})
            
        # Study Modes
        if "study_modes" in ps and isinstance(ps["study_modes"], list):
            info_data.append({"key": "study_modes", "value": ", ".join(ps["study_modes"])})
        if "default_mode" in ps:
            info_data.append({"key": "default_mode", "value": str(ps["default_mode"])})
            
        # Practice Defaults
        mcq = ps.get("mcq", {})
        if isinstance(mcq, dict):
            mcq_pairs = mcq.get("active_pairs", [])
            if mcq_pairs:
                info_data.append({"key": "mcq_active_pairs", "value": ", ".join([f"{p.get('q')}-{p.get('a')}" for p in mcq_pairs if isinstance(p, dict)])})
            if "num_choices" in mcq:
                info_data.append({"key": "mcq_num_choices", "value": str(mcq["num_choices"])})
                
        typing = ps.get("typing", {})
        if isinstance(typing, dict):
            typing_pairs = typing.get("active_pairs", [])
            if typing_pairs:
                info_data.append({"key": "typing_active_pairs", "value": ", ".join([f"{p.get('q')}-{p.get('a')}" for p in typing_pairs if isinstance(p, dict)])})
                
        listening = ps.get("listening", {})
        if isinstance(listening, dict):
            listening_pairs = listening.get("active_pairs", [])
            if listening_pairs:
                info_data.append({"key": "listening_active_pairs", "value": ", ".join([f"{p.get('q')}-{p.get('a')}" for p in listening_pairs if isinstance(p, dict)])})
            if "num_choices" in listening:
                info_data.append({"key": "listening_num_choices", "value": str(listening["num_choices"])})

        # AI Intelligence
        exp_prompt = ai_prompt or ps.get("ai_prompt", "")
        hint_prompt = ai_prompt_hint or ps.get("ai_prompt_hint", "")
        mne_prompt = ai_prompt_mnemonic or ps.get("ai_prompt_mnemonic", "")
        if exp_prompt: info_data.append({"key": "ai_prompt_explanation", "value": str(exp_prompt)})
        if hint_prompt: info_data.append({"key": "ai_prompt_hint", "value": str(hint_prompt)})
        if mne_prompt: info_data.append({"key": "ai_prompt_mnemonic", "value": str(mne_prompt)})
        if "ai_prompts" in ps and ps["ai_prompts"]:
            info_data.append({"key": "ai_prompts", "value": json.dumps(ps["ai_prompts"], ensure_ascii=False)})

        # Audio Pairs & Config
        if "audio_pairs" in ps and ps["audio_pairs"]:
            info_data.append({"key": "audio_pairs", "value": json.dumps(ps["audio_pairs"], ensure_ascii=False)})
        if "front_audio_config" in ps and ps["front_audio_config"]:
            info_data.append({"key": "front_audio_config", "value": json.dumps(ps["front_audio_config"], ensure_ascii=False)})
        if "back_audio_config" in ps and ps["back_audio_config"]:
            info_data.append({"key": "back_audio_config", "value": json.dumps(ps["back_audio_config"], ensure_ascii=False)})

        # Collaboration
        if collaborators:
            collab_strs = []
            for c in collaborators:
                uname = c.get("username") or c.get("email") or ""
                role = c.get("role") or "editor"
                if uname:
                    collab_strs.append(f"{uname}:{role}")
            if collab_strs:
                info_data.append({"key": "collaborators", "value": ", ".join(collab_strs)})

        # Creator Study Defaults (Human-readable rows & JSON)
        study_defaults = {}
        if isinstance(ps, dict):
            if isinstance(ps.get("study_defaults"), dict):
                study_defaults.update(ps["study_defaults"])
            from app.modules.deck.utils import STUDY_SETTINGS_KEYS
            for sk in STUDY_SETTINGS_KEYS:
                if sk not in study_defaults and sk in ps and ps[sk] is not None:
                    study_defaults[sk] = ps[sk]

        if study_defaults:
            info_data.append({"key": "study_defaults", "value": json.dumps(study_defaults, ensure_ascii=False)})
            for sk, sv in study_defaults.items():
                if sv is not None:
                    val_str = str(sv).lower() if isinstance(sv, bool) else str(sv)
                    info_data.append({"key": f"study_{sk}", "value": val_str})

        # Full Practice Settings JSON (Backup / Advanced)
        if ps:
            info_data.append({"key": "practice_settings", "value": json.dumps(ps, ensure_ascii=False)})
                
        df_info = pd.DataFrame(info_data)
        
        # 2. Prepare Data sheet rows
        # Discover all custom keys present in any question's others dict (excluding internal _formulas)
        custom_cols = set()
        for q in cards:
            if hasattr(q, 'others') and q.others and isinstance(q.others, dict):
                for k in q.others.keys():
                    if k not in ("id", "item_id", "order_in_container", "front", "back", "explanation", 
                                 "front_audio_content", "back_audio_content", "front_audio_url", "back_audio_url", 
                                 "front_img", "back_img", "options", "_formulas", "_formula"):
                        custom_cols.add(k)
                        
        custom_cols = sorted(list(custom_cols))

        # Check if the deck or any card has formulas
        has_any_formulas = False
        deck_col_formulas = ps.get("column_formulas") if isinstance(ps, dict) else {}
        if isinstance(deck_col_formulas, dict) and deck_col_formulas:
            has_any_formulas = True
        else:
            for q in cards:
                if hasattr(q, 'others') and isinstance(q.others, dict) and q.others.get("_formulas"):
                    has_any_formulas = True
                    break

        from openpyxl.utils import get_column_letter

        has_options = any(getattr(q, 'options', None) for q in cards)
        export_cols = ["id", "front", "back", "front_audio_content", "back_audio_content", "front_audio_url", "back_audio_url", "front_img", "back_img"]
        if has_options:
            export_cols.extend(["option_a", "option_b", "option_c", "option_d"])
        export_cols.extend(custom_cols)

        col_name_to_letter = {col: get_column_letter(idx) for idx, col in enumerate(export_cols, start=1)}
        
        # Prepare rows for Data and Data_Formula
        rows_data = []
        rows_formula = []

        for idx, q in enumerate(cards, start=1):
            excel_row_num = idx + 1 # Header is row 1
            row = {
                "id": getattr(q, 'id', None),
                "front": getattr(q, 'content', ""),
                "back": getattr(q, 'explanation', ""),
                "front_audio_content": getattr(q, 'front_audio_content', ""),
                "back_audio_content": getattr(q, 'back_audio_content', ""),
                "front_audio_url": getattr(q, 'front_audio_url', ""),
                "back_audio_url": getattr(q, 'back_audio_url', ""),
                "front_img": getattr(q, 'front_img', ""),
                "back_img": getattr(q, 'back_img', "")
            }

            if has_options:
                options = getattr(q, 'options', []) or []
                for i in range(4):
                    opt_content = options[i].get('content', "") if i < len(options) and isinstance(options[i], dict) else ""
                    row[f"option_{chr(97+i)}"] = opt_content

            others = getattr(q, 'others', {})
            card_formulas = {}
            if isinstance(others, dict):
                card_formulas = dict(others.get("_formulas") or {})
                for col in custom_cols:
                    row[col] = others.get(col, "")

            if isinstance(deck_col_formulas, dict):
                for c_col, c_form in deck_col_formulas.items():
                    if c_col not in card_formulas and c_form:
                        card_formulas[c_col] = c_form

            rows_data.append(row)

            if has_any_formulas:
                row_form = dict(row)
                for col_key, tok_form in card_formulas.items():
                    if col_key in col_name_to_letter:
                        row_form[col_key] = transpile_formula(tok_form, excel_row_num, col_name_to_letter)
                rows_formula.append(row_form)
            
        df_data = pd.DataFrame(rows_data)
        
        # Write to Excel
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_info.to_excel(writer, sheet_name="Info", index=False)
            df_data.to_excel(writer, sheet_name="Data", index=False)
            if has_any_formulas:
                df_formula = pd.DataFrame(rows_formula)
                df_formula.to_excel(writer, sheet_name="Data_Formula", index=False)
            
        return output.getvalue()

    @staticmethod
    def generate_template_excel(output_path: Optional[str] = None) -> bytes:
        """
        Generates a standardized Vocaburn Excel Template containing a comprehensive Info sheet
        (with all possible metadata & study settings fields + guides in columns C, D, E)
        and an example Data sheet.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # 1. SHEET: Info
        ws_info = wb.active
        ws_info.title = "Info"

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        section_font = Font(name="Calibri", size=11, bold=True, color="1E1B4B")
        section_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
        key_font = Font(name="Consolas", size=10, bold=True, color="3730A3")
        val_font = Font(name="Calibri", size=10, bold=False, color="111827")
        guide_font = Font(name="Calibri", size=10, italic=False, color="4B5563")
        opt_font = Font(name="Consolas", size=9, color="059669")

        thin_border = Border(
            left=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"),
            bottom=Side(style="thin", color="E5E7EB")
        )

        headers = [
            "Key",
            "Value",
            "Tên Tiếng Việt",
            "Tùy Chọn Hợp Lệ",
            "Hướng Dẫn & Giải Thích Chi Tiết"
        ]

        ws_info.append(headers)
        for col_idx in range(1, 6):
            cell = ws_info.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_info.row_dimensions[1].height = 28

        info_rows = [
            # Section 1
            ("--- THÔNG TIN CHUNG BỘ THẺ (General Information) ---", "", "", "", ""),
            ("title", "500 Từ Vựng N2 Cơ Bản", "Tên bộ thẻ", "Văn bản tự do", "BẮT BUỘC: Tiêu đề hiển thị của bộ thẻ flashcard"),
            ("description", "Bộ thẻ luyện từ vựng N2 tiếng Nhật có audio và ví dụ minh họa", "Mô tả bộ thẻ", "Văn bản tự do", "Mô tả tóm tắt nội dung, đối tượng học và mục tiêu"),
            ("category", "Tiếng Nhật", "Danh mục", "General, Tiếng Nhật, Tiếng Anh, Y Khoa...", "Tên danh mục phân loại bộ thẻ (hệ thống tự tạo nếu chưa có)"),
            ("tags", "JLPT, N2, Từ vựng, Kanji", "Thẻ phân loại", "Các từ khóa cách nhau bằng dấu phẩy (,)", "Thẻ tìm kiếm giúp người dùng dễ dàng lọc và khám phá"),
            ("cover_image", "https://images.unsplash.com/photo-1528164344705-475426879c0d", "URL Ảnh bìa", "URL ảnh hợp lệ (jpg, png, webp)", "Ảnh minh họa bìa hiển thị trên danh sách bộ thẻ"),
            ("instruction", "Hãy đọc to từ vựng trước khi lật mặt sau kiểm tra nghĩa.", "Hướng dẫn học", "Văn bản tự do", "Ghi chú, mẹo hoặc chỉ dẫn của giáo viên dành cho học viên"),
            ("is_public", "TRUE", "Công khai", "TRUE | FALSE", "TRUE: Mọi người đều có thể tìm và học. FALSE: Chỉ riêng bạn xem được"),
            ("time_limit", "0", "Giới hạn thời gian", "Số nguyên (phút), 0 = Không giới hạn", "Thời gian làm bài tối đa khi người học luyện tập bộ thẻ này"),

            # Section 2
            ("--- CÀI ĐẶT HỌC MẶC ĐỊNH ĐẦU VÀO (Creator Study Defaults) ---", "", "", "", ""),
            ("study_autoplay_audio", "front", "Tự động phát âm thanh", "none | front | back | always", "Tự động phát TTS/Audio: none (tắt), front (mặt trước), back (mặt sau), always (cả hai)"),
            ("study_show_images", "always", "Hiển thị hình ảnh", "always | front | back | none", "Chế độ ảnh: always (luôn hiện), front (chỉ mặt trước), back (chỉ mặt sau), none (ẩn ảnh)"),
            ("study_learning_mode", "fsrs", "Chế độ học mặc định", "fsrs | roadmap | new | review | hardest | flip", "Chế độ khởi đầu khi người học bấm học thẻ: fsrs (giãn cách), roadmap (lộ trình), flip (lật nhanh)..."),
            ("study_random_enabled", "FALSE", "Xáo trộn thứ tự thẻ", "TRUE | FALSE", "TRUE: Ngẫu nhiên thứ tự thẻ khi bắt đầu học; FALSE: Theo thứ tự gốc trong bảng"),
            ("study_sfx_enabled", "TRUE", "Âm thanh hiệu ứng", "TRUE | FALSE", "TRUE: Bật âm thanh chúc mừng / âm thanh phản hồi thao tác"),
            ("study_quick_learn_enabled", "FALSE", "Chế độ học nhanh", "TRUE | FALSE", "TRUE: Tự động chuyển thẻ kế tiếp ngay khi người học chọn hoặc đánh giá"),
            ("study_haptic_enabled", "TRUE", "Rung phản hồi (Điện thoại)", "TRUE | FALSE", "TRUE: Rung nhẹ thiết bị khi thao tác trên ứng dụng di động"),
            ("study_show_fsrs", "TRUE", "Nút đánh giá FSRS", "TRUE | FALSE", "TRUE: Hiển thị 4 nút đánh giá FSRS (Again, Hard, Good, Easy)"),

            # Section 3
            ("--- CẤU HÌNH LUYỆN TẬP ĐA CHẾ ĐỘ (Practice Modes & Pairs) ---", "", "", "", ""),
            ("mcq_active_pairs", "front-back, back-front", "Cặp luyện tập trắc nghiệm", "Cặp tên cột dạng: q-a, q2-a2", "Cột câu hỏi và cột đáp án tạo đề trắc nghiệm 4 đáp án"),
            ("mcq_num_choices", "4", "Số lựa chọn trắc nghiệm", "Từ 3 đến 8 (mặc định 4)", "Số lượng đáp án A, B, C, D... hiển thị cho mỗi câu hỏi"),
            ("typing_active_pairs", "back-front", "Cặp luyện tập gõ từ", "Cặp tên cột dạng: q-a", "Hiển thị cột q (gợi ý nghĩa) và yêu cầu người học gõ chính xác cột a"),
            ("listening_active_pairs", "front_audio-back", "Cặp luyện tập nghe hiểu", "Cặp tên cột dạng: q-a", "Chế độ nghe audio phát ra và bấm chọn đáp án đúng"),
            ("listening_num_choices", "4", "Số lựa chọn bài nghe", "Từ 3 đến 8 (mặc định 4)", "Số lựa chọn đáp án hiển thị trong bài luyện tập nghe"),

            # Section 4
            ("--- CẤU HÌNH TRÍ TUỆ NHÂN TẠO AI (AI Prompts) ---", "", "", "", ""),
            ("ai_prompt_explanation", "Hãy giải thích chi tiết ý nghĩa và ngữ cảnh sử dụng của từ này.", "Prompt giải thích từ vựng", "Câu lệnh AI tự do", "Lời nhắc gửi tới AI khi người học bấm nút 'Giải thích bằng AI'"),
            ("ai_prompt_hint", "Hãy tạo 1 câu đố ngắn không chứa từ khóa để người học tự đoán.", "Prompt gợi ý (Hint)", "Câu lệnh AI tự do", "Lời nhắc gửi tới AI để sinh gợi ý khi người học gặp khó"),
            ("ai_prompt_mnemonic", "Hãy tạo mẹo ghi nhớ vui nhộn hoặc chiết tự chữ Hán cho từ này.", "Prompt mẹo ghi nhớ", "Câu lệnh AI tự do", "Lời nhắc gửi tới AI để tạo câu chuyện ghi nhớ từ vựng"),

            # Section 5
            ("--- QUẢN LÝ CỘNG TÁC VIÊN (Collaborators) ---", "", "", "", ""),
            ("collaborators", "teacher_minh:editor, trogiang_nam:editor", "Cộng tác viên biên tập", "username:role (editor | viewer)", "Danh sách người dùng được quyền xem hoặc cùng sửa bộ thẻ, cách nhau dấu phẩy"),

            # Section 6
            ("--- CẤU HÌNH NÂNG CAO (Advanced JSON Backup) ---", "", "", "", ""),
            ("study_defaults", "{}", "JSON cấu hình học mặc định", "Chuỗi JSON hợp lệ", "Chuỗi JSON chứa tất cả study_defaults (tự động cập nhật nếu nhập lẻ bên trên)"),
            ("practice_settings", "{}", "JSON toàn bộ practice settings", "Chuỗi JSON hợp lệ", "Chuỗi JSON sao lưu toàn diện tất cả các tab cấu hình bộ thẻ")
        ]

        current_row = 2
        for item in info_rows:
            key, val, name_vn, options, guide = item
            if key.startswith("---"):
                ws_info.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                c = ws_info.cell(row=current_row, column=1, value=key)
                c.font = section_font
                c.fill = section_fill
                c.alignment = Alignment(horizontal="left", vertical="center")
                ws_info.row_dimensions[current_row].height = 22
            else:
                c1 = ws_info.cell(row=current_row, column=1, value=key)
                c1.font = key_font
                c1.border = thin_border
                c1.alignment = Alignment(horizontal="left", vertical="center")

                c2 = ws_info.cell(row=current_row, column=2, value=val)
                c2.font = val_font
                c2.border = thin_border
                c2.alignment = Alignment(horizontal="left", vertical="center")

                c3 = ws_info.cell(row=current_row, column=3, value=name_vn)
                c3.font = guide_font
                c3.border = thin_border
                c3.alignment = Alignment(horizontal="left", vertical="center")

                c4 = ws_info.cell(row=current_row, column=4, value=options)
                c4.font = opt_font
                c4.border = thin_border
                c4.alignment = Alignment(horizontal="left", vertical="center")

                c5 = ws_info.cell(row=current_row, column=5, value=guide)
                c5.font = guide_font
                c5.border = thin_border
                c5.alignment = Alignment(horizontal="left", vertical="center")

                ws_info.row_dimensions[current_row].height = 20
            current_row += 1

        ws_info.column_dimensions["A"].width = 28
        ws_info.column_dimensions["B"].width = 24
        ws_info.column_dimensions["C"].width = 28
        ws_info.column_dimensions["D"].width = 36
        ws_info.column_dimensions["E"].width = 65

        # 2. SHEET: Data
        ws_data = wb.create_sheet(title="Data")
        data_header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")

        data_headers = [
            "front",
            "back",
            "explanation",
            "front_audio_content",
            "back_audio_content",
            "front_audio_url",
            "back_audio_url",
            "front_img",
            "back_img",
            "kanji",
            "furigana",
            "example"
        ]

        ws_data.append(data_headers)
        for col_idx in range(1, len(data_headers) + 1):
            cell = ws_data.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = data_header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_data.row_dimensions[1].height = 28

        sample_rows = [
            [
                "こんにちは",
                "Xin chào",
                "Lời chào thông dụng buổi sáng/chiều trong giao tiếp tiếng Nhật.",
                "こんにちは",
                "Xin chào",
                "",
                "",
                "https://images.unsplash.com/photo-1528164344705-475426879c0d",
                "",
                "今日",
                "konnichiwa",
                "皆さん、こんにちは！"
            ],
            [
                "ありがとう",
                "Cảm ơn",
                "Lời cảm ơn lịch sự cơ bản.",
                "ありがとう",
                "Cảm ơn",
                "",
                "",
                "",
                "",
                "有難う",
                "arigatou",
                "どうもありがとうございます。"
            ],
            [
                "勉強 (べんきょう)",
                "Học tập, học hỏi",
                "Danh từ hoặc động từ Suru: học tập chuyên cần.",
                "べんきょう",
                "Học tập",
                "",
                "",
                "",
                "",
                "勉強",
                "benkyou",
                "毎日日本語を勉強しています。"
            ]
        ]

        for r_idx, row_vals in enumerate(sample_rows, start=2):
            for c_idx, val in enumerate(row_vals, start=1):
                cell = ws_data.cell(row=r_idx, column=c_idx, value=val)
                cell.font = val_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
            ws_data.row_dimensions[r_idx].height = 20

        for col_idx, h in enumerate(data_headers, start=1):
            col_letter = get_column_letter(col_idx)
            ws_data.column_dimensions[col_letter].width = max(len(h) + 6, 20)

        out_stream = BytesIO()
        wb.save(out_stream)
        tmpl_bytes = out_stream.getvalue()

        if output_path:
            with open(output_path, "wb") as f:
                f.write(tmpl_bytes)

        return tmpl_bytes
